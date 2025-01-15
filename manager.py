import time
import openpyxl
import sulkusers_data.creador as data 
import sulkusers_credits.creaUsuario as credits

#¿Cuantos?:
usuarios = 1
creditos = 11
aplicacion = "superheroes"
ambiente = "prod"

archivo_excel = "usuarios-ronda1.xlsx"

#Especificaciones Excel: 
# Comprobar si el archivo existe
try:
    workbook = openpyxl.load_workbook(archivo_excel)
    sheet = workbook.active

    #Estilo
    # Crear un estilo de fuente en negrita
    bold_font = openpyxl.styles.Font(bold=True)
    # Aplicar el estilo a la fila de los encabezados
    for cell in sheet[1:1]:  # Selecciona todas las celdas de la primera fila
        cell.font = bold_font
    # Ajustar el ancho de la columna A a 20 unidades
    sheet.column_dimensions['A'].auto_size = True
    sheet.column_dimensions['B'].auto_size = True
    sheet.column_dimensions['C'].auto_size = True
except FileNotFoundError:
    # Si no existe, crear un nuevo libro y hoja
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Usuarios"
    # Agregar los encabezados (si es la primera vez)
    sheet['A1'] = 'Usuario'
    sheet['B1'] = 'Contraseña'
    sheet['C1'] = 'Créditos'

# Obtener la última fila con datos
last_row = sheet.max_row

for i in range(usuarios):
    #Alta de Nuevo Usuario dentro de Sulkusers_Data
    #Solo aplicación si es un superheroe-random, aplicacion, usuario y contraseña si es un usuario específico.
    usuario_creado, contrasena = data.nuevoUsuario(aplicacion, "", "")
    print(f"Usuario {usuario_creado} listo.")
    time.sleep(1)
    #Alta o Recarga de Créditos dentro de Sulkusers_Credits
    #Aquí le tendrás que decir el usuario que creaste previamente en sulkusers_data
    credits.crear_usuario(usuario_creado, ambiente, creditos)

    sheet.append([usuario_creado, contrasena, creditos])

    #Guardamos los 3 datos en el excel: 
    # Escribir los datos en la siguiente fila
    # fila = i + 2  # Comenzamos en la fila 2 para dejar la fila 1 para los encabezados
    # sheet.cell(row=fila, column=1, value=usuario_creado)
    # sheet.cell(row=fila, column=2, value=contrasena)
    # sheet.cell(row=fila, column=3, value=creditos)



# Guardar el archivo de Excel
workbook.save(archivo_excel)